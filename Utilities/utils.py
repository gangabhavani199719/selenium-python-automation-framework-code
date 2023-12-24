import logging
import inspect
import softest
from openpyxl import Workbook,load_workbook
import csv
class Utils(softest.TestCase):
    def assertListItemText(self,list,value):

        for stop in list :
            print("the text is:" + stop.text)
            self.soft_assert(self.assertEqual,stop.text,value)
            print("assert pass")
            if stop.text == value:
                print("test passed")
            else:
                print("test failed")
        self.assert_all()

    def  custom_logger(logLevel=logging.DEBUG):
        # set class/method name from where its called
        logger_name=inspect.stack()[1][3]
        #create logger
        logger=logging.getLogger(logger_name)
        logger.setLevel(logLevel)
        # create console handler or file handler and set the log level
        fh = logging.FileHandler("automation.log",mode='a')  # by default its append mode or u can update as override mode as 'w'
        # create formatter - how you want your logs to be formatted
        formatter =logging.Formatter('%(asctime)s - %(levelname)s : %(message)s',datefmt='%m/%d/%Y %I:%M:%S %p')
        # add formatter to conole or file handler
        fh.setFormatter(formatter)
        # add console handler to logger
        logger.addHandler(fh)
        return logger
    def read_data_from_excel(file_name,sheet):
        datalist=[]
        wb=load_workbook(filename=file_name)
        sh=wb[sheet]
        row_ct=sh.max_row
        col_ct=sh.max_column

        for i in range(2,row_ct+1):
            row=[]
            for j in range(1,col_ct+1):
                row.append(sh.cell(row=i,column=j).value)
            datalist.append(row)
        return datalist

    def read_from_csv_data(file_name):
        #create an empty list
        datalist=[]
        #open CSV file
        csvdata=open(file_name,"r")
        #create csv header
        reader=csv.reader(csvdata)
        #skip header
        next(reader)
        #Add csv rows to list
        for rows in reader:
            datalist.append(rows)
        return datalist
